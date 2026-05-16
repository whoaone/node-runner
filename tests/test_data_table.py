"""v5.2.0 item 45: tests for the Data Table dialog.

We don't need a full MainWindow -- the dialog talks to MW via:
- ``selection_changed`` signal (str, list)
- ``op2_results`` attribute (dict)
- ``_current_selection_nids`` / ``_current_selection_eids`` attributes

So we fake a tiny MainWindow stub with those four members.
"""

from __future__ import annotations

import csv
import os
import tempfile

import pytest
from PySide6.QtCore import Signal
from PySide6.QtWidgets import QWidget


class FakeMainWindow(QWidget):
    """Minimal stub exposing the same surface the dialog needs."""

    selection_changed = Signal(str, list)

    def __init__(self):
        super().__init__()
        self.op2_results = {
            'filepath': 'fake.f06',
            'results_source': 'f06',
            'subcases': {
                1: {
                    'title': 'Static',
                    'displacements': {
                        1: [0.1, 0.0, 0.0, 0.0, 0.0, 0.0],
                        2: [0.0, 0.2, 0.0, 0.0, 0.0, 0.0],
                        3: [0.0, 0.0, 0.3, 0.0, 0.0, 0.0],
                    },
                    'stresses': {
                        100: {'von_mises': 1000.0, 'oxx': 100.0, 'oyy': 0.0,
                              'txy': 0.0, 'max_principal': 100.0,
                              'min_principal': 0.0},
                        200: {'von_mises': 2000.0, 'oxx': 200.0, 'oyy': 0.0,
                              'txy': 0.0, 'max_principal': 200.0,
                              'min_principal': 0.0},
                    },
                },
            },
        }
        self._current_selection_nids = [1, 2, 3]
        self._current_selection_eids = [100, 200]


def test_dialog_constructs(qtbot):
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    # Output Set combo populated with one entry.
    assert dlg._output_set_combo.count() == 1
    # Source = Nodes by default; status reads "3 nodes" with 0 columns.
    assert "3 node" in dlg._status_lbl.text().lower()


def test_live_sync_emits_rebuild(qtbot):
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    # Change the selection to a single node and emit the signal.
    mw._current_selection_nids = [2]
    mw.selection_changed.emit('Node', [2])
    assert "1 node" in dlg._status_lbl.text().lower()


def test_add_displacement_column(qtbot):
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    # Programmatically add a Disp T1 column (skip the dialog popup).
    cols = dlg._available_columns()
    t1_col = next(c for c in cols if c[3] == 'Disp T1')
    dlg._columns.append(t1_col)
    dlg._rebuild_table()
    model = dlg._view.model()
    # 3 rows (nids 1,2,3) x 2 cols (NID + Disp T1)
    assert model.rowCount() == 3
    assert model.columnCount() == 2
    # Row 0 corresponds to NID 1, whose T1 = 0.1
    assert model.data(model.index(0, 1)) == "0.1"


def test_source_toggle_to_elements(qtbot):
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    dlg._radio_elems.setChecked(True)
    assert dlg._source_kind == 'Element'
    # 2 elements selected, no columns yet -> 2 rows x 1 col (EID)
    model = dlg._view.model()
    assert model.rowCount() == 2
    # Now add stress column.
    cols = dlg._available_columns()
    vm_col = next(c for c in cols if 'vM' in c[3])
    dlg._columns.append(vm_col)
    dlg._rebuild_table()
    model = dlg._view.model()
    assert model.columnCount() == 2
    # First element EID=100 -> von_mises = 1000
    assert model.data(model.index(0, 1)) == "1000"


def test_csv_export_creates_valid_file(qtbot, tmp_path):
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    cols = dlg._available_columns()
    dlg._columns.append(next(c for c in cols if c[3] == 'Disp T1'))
    dlg._rebuild_table()
    headers, rows = dlg._gather_rows()
    csv_path = tmp_path / "out.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
    text = csv_path.read_text(encoding='utf-8')
    assert "NID,Disp T1" in text or "NID,\"Disp T1\"" in text
    assert "0.1" in text


def test_live_sync_survives_close_and_reopen(qtbot):
    """Regression for v5.2.0 Round-2 Bug #3: closeEvent used to
    disconnect ``selection_changed``, silently breaking live-sync on
    the second open."""
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    dlg.show()
    dlg.close()
    dlg.show()    # re-open
    # Emit a selection change AFTER the close/show cycle.
    mw._current_selection_nids = [3]
    mw.selection_changed.emit('Node', [3])
    # The dialog must have rebuilt with 1 row, not still be at 3 rows.
    assert "1 node" in dlg._status_lbl.text().lower()


def test_xlsx_export_creates_valid_file(qtbot, tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    from node_runner.dialogs.data_table import DataTableDialog
    mw = FakeMainWindow()
    qtbot.addWidget(mw)
    dlg = DataTableDialog(mw)
    qtbot.addWidget(dlg)
    cols = dlg._available_columns()
    dlg._columns.append(next(c for c in cols if c[3] == 'Disp T1'))
    dlg._rebuild_table()
    headers, rows = dlg._gather_rows()
    xlsx_path = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([str(h) for h in headers])
    for row in rows:
        ws.append([v for v in row])
    wb.save(xlsx_path)
    # Re-open and assert the contents survived.
    wb2 = openpyxl.load_workbook(xlsx_path)
    ws2 = wb2.active
    assert ws2.cell(1, 1).value == "NID"
    assert ws2.cell(1, 2).value == "Disp T1"
